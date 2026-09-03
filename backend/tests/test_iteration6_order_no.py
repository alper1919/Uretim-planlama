"""Iteration 6: order_no field CRUD + Excel export column + regression file cookie auth."""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

def _load_frontend_env():
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                return line.split("=", 1)[1].strip()
    raise RuntimeError("REACT_APP_BACKEND_URL not found")

BASE_URL = (os.environ.get("REACT_APP_BACKEND_URL") or _load_frontend_env()).rstrip("/")
TOKEN = "seed_session_alper"
H = {"Authorization": f"Bearer {TOKEN}"}


@pytest.fixture(scope="module")
def created_part_id():
    """Create part with order_no; cleanup after."""
    payload = {
        "order_no": "TEST_SIP-2026-999",
        "part_code": "TEST_PRC_ITER6",
        "part_name": "Iter6 Test Part",
        "quantity": 3,
        "priority": "yuksek",
    }
    r = requests.post(f"{BASE_URL}/api/parts", json=payload, headers=H)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["order_no"] == "TEST_SIP-2026-999"
    pid = data["id"]
    yield pid
    requests.delete(f"{BASE_URL}/api/parts/{pid}", headers=H)


class TestOrderNoCRUD:
    def test_create_persists_order_no(self, created_part_id):
        r = requests.get(f"{BASE_URL}/api/parts/{created_part_id}", headers=H)
        assert r.status_code == 200
        d = r.json()
        assert d["order_no"] == "TEST_SIP-2026-999"
        assert d["part_code"] == "TEST_PRC_ITER6"

    def test_list_includes_order_no(self, created_part_id):
        r = requests.get(f"{BASE_URL}/api/parts", headers=H)
        assert r.status_code == 200
        parts = r.json()
        found = [p for p in parts if p["id"] == created_part_id]
        assert len(found) == 1
        assert found[0]["order_no"] == "TEST_SIP-2026-999"

    def test_update_order_no(self, created_part_id):
        payload = {
            "order_no": "TEST_SIP-2026-UPDATED",
            "part_code": "TEST_PRC_ITER6",
            "part_name": "Iter6 Test Part",
            "quantity": 3,
            "priority": "yuksek",
        }
        r = requests.put(f"{BASE_URL}/api/parts/{created_part_id}", json=payload, headers=H)
        assert r.status_code == 200
        assert r.json()["order_no"] == "TEST_SIP-2026-UPDATED"
        # Verify GET persists
        r2 = requests.get(f"{BASE_URL}/api/parts/{created_part_id}", headers=H)
        assert r2.json()["order_no"] == "TEST_SIP-2026-UPDATED"

    def test_create_without_order_no_defaults_empty(self):
        payload = {"part_code": "TEST_PRC_NOORDER", "part_name": "No Order"}
        r = requests.post(f"{BASE_URL}/api/parts", json=payload, headers=H)
        assert r.status_code == 200
        d = r.json()
        assert d["order_no"] == ""
        # cleanup
        requests.delete(f"{BASE_URL}/api/parts/{d['id']}", headers=H)


class TestExcelExport:
    def test_export_returns_xlsx_with_siparis_no_column(self, created_part_id):
        r = requests.get(f"{BASE_URL}/api/export/parts.xlsx", headers=H)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Parçalar"]
        headers = [c.value for c in ws[1]]
        assert headers[0] == "Sipariş No", f"First col should be Sipariş No, got {headers}"
        # Find our test part's row and confirm order_no value in col A
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == "TEST_PRC_ITER6":
                assert row[0] in ("TEST_SIP-2026-UPDATED", "TEST_SIP-2026-999")
                found = True
                break
        assert found, "TEST part not present in export"


class TestFileCookieAuthRegression:
    """Regression: file endpoint accepts cookie, bearer, ?auth."""

    @pytest.fixture(scope="class")
    def uploaded_pdf(self, created_part_id):
        pdf = b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 100 100]>>endobj\nxref\n0 4\n0000000000 65535 f\n0000000010 00000 n\n0000000053 00000 n\n0000000100 00000 n\ntrailer<</Size 4/Root 1 0 R>>\nstartxref\n160\n%%EOF"
        r = requests.post(
            f"{BASE_URL}/api/parts/{created_part_id}/drawings",
            headers=H,
            files={"file": ("iter6.pdf", io.BytesIO(pdf), "application/pdf")},
        )
        assert r.status_code == 200
        d = [x for x in r.json()["drawings"] if x["original_filename"] == "iter6.pdf"][-1]
        return d["storage_path"]

    def test_cookie_auth(self, uploaded_pdf):
        r = requests.get(f"{BASE_URL}/api/files/{uploaded_pdf}", cookies={"session_token": TOKEN})
        assert r.status_code == 200
        assert r.headers.get("content-type") == "application/pdf"
        assert r.content.startswith(b"%PDF")

    def test_bearer_auth(self, uploaded_pdf):
        r = requests.get(f"{BASE_URL}/api/files/{uploaded_pdf}", headers=H)
        assert r.status_code == 200

    def test_query_auth(self, uploaded_pdf):
        r = requests.get(f"{BASE_URL}/api/files/{uploaded_pdf}?auth={TOKEN}")
        assert r.status_code == 200

    def test_no_auth_401(self, uploaded_pdf):
        r = requests.get(f"{BASE_URL}/api/files/{uploaded_pdf}")
        assert r.status_code == 401
