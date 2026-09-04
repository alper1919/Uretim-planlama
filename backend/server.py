from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Header, Query, Response, Request
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import uuid
import requests
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------------- Object Storage ----------------
STORAGE_BASE = (os.environ.get("INTEGRATION_PROXY_URL") or "").strip() or "https://integrations.emergentagent.com"
STORAGE_URL = STORAGE_BASE.rstrip("/") + "/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "parcatakip"
storage_key = None

MIME_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "gif": "image/gif",
    "webp": "image/webp", "pdf": "application/pdf", "dwg": "application/acad",
    "dxf": "image/vnd.dxf", "step": "application/step", "stp": "application/step",
    "igs": "model/iges", "iges": "model/iges",
}


def init_storage(force: bool = False):
    global storage_key
    if storage_key and not force:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key


def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(f"{STORAGE_URL}/objects/{path}",
                        headers={"X-Storage-Key": key, "Content-Type": content_type},
                        data=data, timeout=120)
    if resp.status_code == 404:
        key = init_storage(force=True)
        resp = requests.put(f"{STORAGE_URL}/objects/{path}",
                            headers={"X-Storage-Key": key, "Content-Type": content_type},
                            data=data, timeout=120)
    resp.raise_for_status()
    return resp.json()


def get_object(path: str):
    key = init_storage()
    resp = requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    if resp.status_code == 404:
        key = init_storage(force=True)
        resp = requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")


# ---------------- Models ----------------
STATUSES = [
    "hammadde_siparis_edildi", "hammadde_geldi", "isleme_alindi", "uretim_bitti",
    "tesviyede", "kalite_kontrol", "sevk_alaninda", "sevk_edildi",
]

STATUS_LABELS = {
    "hammadde_siparis_edildi": "Hammadde Sipariş Edildi",
    "hammadde_geldi": "Hammadde Geldi",
    "isleme_alindi": "İşleme Alındı",
    "uretim_bitti": "Üretim Bitti",
    "tesviyede": "Tesviyede",
    "kalite_kontrol": "Kalite Kontrol",
    "sevk_alaninda": "Sevk Alanında",
    "sevk_edildi": "Sevk Edildi",
}

PRIORITY_LABELS = {"dusuk": "Düşük", "normal": "Normal", "yuksek": "Yüksek", "acil": "Acil"}


class User(BaseModel):
    user_id: str
    username: str
    name: str
    role: str = "user"


class DrawingFile(BaseModel):
    id: str
    storage_path: str
    original_filename: str
    content_type: str
    size: int
    uploaded_at: str


class HistoryEntry(BaseModel):
    id: str
    from_status: Optional[str] = None
    to_status: str
    note: Optional[str] = None
    user_name: str
    user_picture: Optional[str] = None
    timestamp: str


class PartCreate(BaseModel):
    order_no: str = ""
    part_code: str
    part_name: str
    quantity: int = 1
    material_type: str = ""
    material_dimensions: str = ""
    priority: str = "normal"
    workstation: str = ""
    customer: str = ""
    notes: str = ""
    due_date: Optional[str] = None


class Part(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_no: str = ""
    part_code: str
    part_name: str
    quantity: int = 1
    material_type: str = ""
    material_dimensions: str = ""
    priority: str = "normal"
    workstation: str = ""
    customer: str = ""
    notes: str = ""
    due_date: Optional[str] = None
    status: str = "hammadde_siparis_edildi"
    drawings: List[DrawingFile] = []
    history: List[HistoryEntry] = []
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class StatusUpdate(BaseModel):
    status: str
    note: Optional[str] = None


# ---------------- Auth (JWT username/password) ----------------
JWT_ALGORITHM = "HS256"


def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode("utf-8"), h.encode("utf-8"))
    except Exception:
        return False


def create_token(user_id: str, days: int) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=days), "type": "access"}
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGORITHM)


def to_user(doc: dict) -> User:
    return User(**{k: doc.get(k) for k in ("user_id", "username", "name", "role")})


class LoginRequest(BaseModel):
    username: str
    password: str
    remember: bool = False


class UserCreate(BaseModel):
    username: str
    password: str
    name: str = ""
    role: str = "user"


async def get_current_user(request: Request, authorization: Optional[str] = Header(None)) -> User:
    token = request.cookies.get("access_token")
    if not token and authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ", 1)[1]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user_doc = await db.users.find_one({"user_id": payload.get("sub")}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    return to_user(user_doc)


async def require_admin(user: User = Depends(get_current_user)) -> User:
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için admin yetkisi gerekli")
    return user


@api_router.post("/auth/login", response_model=User)
async def login(payload: LoginRequest, response: Response):
    username = payload.username.strip().lower()
    user_doc = await db.users.find_one({"username": username}, {"_id": 0})
    if not user_doc or not verify_password(payload.password, user_doc.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Kullanıcı adı veya şifre hatalı")
    days = 30 if payload.remember else 1
    token = create_token(user_doc["user_id"], days)
    kwargs = dict(key="access_token", value=token, httponly=True, secure=True, samesite="none", path="/")
    if payload.remember:
        kwargs["max_age"] = days * 24 * 60 * 60
    response.set_cookie(**kwargs)
    return to_user(user_doc)


@api_router.get("/auth/me", response_model=User)
async def auth_me(user: User = Depends(get_current_user)):
    return user


@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api_router.get("/users", response_model=List[User])
async def list_users(admin: User = Depends(require_admin)):
    docs = await db.users.find({}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    return [to_user(d) for d in docs]


@api_router.post("/users", response_model=User)
async def create_user(payload: UserCreate, admin: User = Depends(require_admin)):
    username = payload.username.strip().lower()
    if not username or not payload.password:
        raise HTTPException(status_code=400, detail="Kullanıcı adı ve şifre zorunludur")
    if await db.users.find_one({"username": username}):
        raise HTTPException(status_code=400, detail="Bu kullanıcı adı zaten kayıtlı")
    role = "admin" if payload.role == "admin" else "user"
    doc = {"user_id": f"user_{uuid.uuid4().hex[:12]}", "username": username,
           "password_hash": hash_password(payload.password), "name": payload.name.strip() or username,
           "role": role, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.users.insert_one(doc)
    return to_user(doc)


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: User = Depends(require_admin)):
    if user_id == admin.user_id:
        raise HTTPException(status_code=400, detail="Kendinizi silemezsiniz")
    res = await db.users.delete_one({"user_id": user_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    return {"ok": True}


# ---------------- Parts ----------------
def now_iso():
    return datetime.now(timezone.utc).isoformat()


@api_router.get("/parts", response_model=List[Part])
async def list_parts(user: User = Depends(get_current_user)):
    docs = await db.parts.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [Part(**d) for d in docs]


@api_router.post("/parts", response_model=Part)
async def create_part(payload: PartCreate, user: User = Depends(get_current_user)):
    part = Part(**payload.model_dump(), created_by=user.name)
    part.history = [HistoryEntry(
        id=str(uuid.uuid4()), from_status=None, to_status=part.status,
        note="Parça oluşturuldu", user_name=user.name, user_picture=None, timestamp=now_iso(),
    )]
    await db.parts.insert_one(part.model_dump())
    return part


@api_router.get("/parts/{part_id}", response_model=Part)
async def get_part(part_id: str, user: User = Depends(get_current_user)):
    doc = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    return Part(**doc)


@api_router.put("/parts/{part_id}", response_model=Part)
async def update_part(part_id: str, payload: PartCreate, user: User = Depends(get_current_user)):
    doc = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    update = payload.model_dump()
    update["updated_at"] = now_iso()
    await db.parts.update_one({"id": part_id}, {"$set": update})
    doc = await db.parts.find_one({"id": part_id}, {"_id": 0})
    return Part(**doc)


@api_router.delete("/parts/{part_id}")
async def delete_part(part_id: str, user: User = Depends(get_current_user)):
    res = await db.parts.delete_one({"id": part_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    return {"ok": True}


@api_router.patch("/parts/{part_id}/status", response_model=Part)
async def update_status(part_id: str, payload: StatusUpdate, user: User = Depends(get_current_user)):
    if payload.status not in STATUSES:
        raise HTTPException(status_code=400, detail="Geçersiz durum")
    doc = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    entry = HistoryEntry(
        id=str(uuid.uuid4()), from_status=doc["status"], to_status=payload.status,
        note=payload.note, user_name=user.name, user_picture=None, timestamp=now_iso(),
    )
    await db.parts.update_one({"id": part_id}, {
        "$set": {"status": payload.status, "updated_at": now_iso()},
        "$push": {"history": entry.model_dump()},
    })
    doc = await db.parts.find_one({"id": part_id}, {"_id": 0})
    return Part(**doc)


# ---------------- Files ----------------
@api_router.post("/parts/{part_id}/drawings", response_model=Part)
async def upload_drawing(part_id: str, file: UploadFile = File(...), user: User = Depends(get_current_user)):
    doc = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "bin"
    content_type = MIME_TYPES.get(ext, file.content_type or "application/octet-stream")
    path = f"{APP_NAME}/drawings/{part_id}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    result = put_object(path, data, content_type)
    drawing = DrawingFile(
        id=str(uuid.uuid4()), storage_path=result["path"], original_filename=file.filename,
        content_type=content_type, size=result.get("size", len(data)), uploaded_at=now_iso(),
    )
    await db.parts.update_one({"id": part_id}, {"$push": {"drawings": drawing.model_dump()},
                                                "$set": {"updated_at": now_iso()}})
    doc = await db.parts.find_one({"id": part_id}, {"_id": 0})
    return Part(**doc)


@api_router.delete("/parts/{part_id}/drawings/{drawing_id}", response_model=Part)
async def delete_drawing(part_id: str, drawing_id: str, user: User = Depends(get_current_user)):
    await db.parts.update_one({"id": part_id}, {"$pull": {"drawings": {"id": drawing_id}},
                                                "$set": {"updated_at": now_iso()}})
    doc = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    return Part(**doc)


@api_router.get("/files/{path:path}")
async def download_file(path: str, request: Request, authorization: Optional[str] = Header(None), auth: Optional[str] = Query(None)):
    token = request.cookies.get("access_token")
    if not token and authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ", 1)[1]
    if not token and auth:
        token = auth
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    part = await db.parts.find_one({"drawings.storage_path": path}, {"_id": 0})
    content_type = "application/octet-stream"
    filename = "file"
    if part:
        for d in part["drawings"]:
            if d["storage_path"] == path:
                content_type = d["content_type"]
                filename = d["original_filename"]
    data, ct = get_object(path)
    return Response(content=data, media_type=content_type,
                    headers={"Content-Disposition": f'inline; filename="{filename}"'})


@api_router.get("/export/parts.xlsx")
async def export_parts(user: User = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    def fmt_dt(iso):
        if not iso:
            return ""
        try:
            return datetime.fromisoformat(iso).strftime("%d.%m.%Y %H:%M")
        except Exception:
            return iso

    docs = await db.parts.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    wb = Workbook()
    header_font = Font(bold=True, color="0B0F17")
    header_fill = PatternFill("solid", fgColor="F59E0B")

    ws = wb.active
    ws.title = "Parçalar"
    headers = ["Sipariş No", "Parça Kodu", "Parça Adı", "Adet", "Hammadde Cinsi", "Hammadde Ölçüleri",
               "Aciliyet", "Durum", "Tezgah", "Müşteri", "Teslim Tarihi", "Teknik Resim",
               "Oluşturan", "Oluşturulma"]
    ws.append(headers)
    for d in docs:
        ws.append([
            d.get("order_no", ""), d.get("part_code", ""), d.get("part_name", ""), d.get("quantity", 0),
            d.get("material_type", ""), d.get("material_dimensions", ""),
            PRIORITY_LABELS.get(d.get("priority"), d.get("priority", "")),
            STATUS_LABELS.get(d.get("status"), d.get("status", "")),
            d.get("workstation", ""), d.get("customer", ""), d.get("due_date") or "",
            len(d.get("drawings", [])), d.get("created_by", ""), fmt_dt(d.get("created_at")),
        ])

    ws2 = wb.create_sheet("Durum Geçmişi")
    ws2.append(["Parça Kodu", "Parça Adı", "Önceki Durum", "Yeni Durum", "Not", "Kullanıcı", "Zaman"])
    for d in docs:
        for h in d.get("history", []):
            ws2.append([
                d.get("part_code", ""), d.get("part_name", ""),
                STATUS_LABELS.get(h.get("from_status"), h.get("from_status") or "—"),
                STATUS_LABELS.get(h.get("to_status"), h.get("to_status", "")),
                h.get("note") or "", h.get("user_name", ""), fmt_dt(h.get("timestamp")),
            ])

    for sheet in (ws, ws2):
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max(width + 3, 12), 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="parcatakip_rapor_{stamp}.xlsx"'},
    )


@api_router.get("/")
async def root():
    return {"message": "ParçaTakip PRO API"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    try:
        init_storage()
        logger.info("Storage initialized")
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
    try:
        await db.users.create_index("username", unique=True)
        admin_username = os.environ.get("ADMIN_USERNAME", "admin").strip().lower()
        admin_password = os.environ["ADMIN_PASSWORD"]
        admin_email = os.environ.get("ADMIN_EMAIL", "")
        existing = await db.users.find_one({"username": admin_username})
        if existing is None:
            await db.users.insert_one({
                "user_id": f"user_{uuid.uuid4().hex[:12]}", "username": admin_username,
                "password_hash": hash_password(admin_password), "name": "Yönetici",
                "role": "admin", "email": admin_email,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            logger.info("Admin user seeded")
        elif not verify_password(admin_password, existing.get("password_hash", "")):
            await db.users.update_one({"username": admin_username},
                                      {"$set": {"password_hash": hash_password(admin_password), "role": "admin"}})
            logger.info("Admin password updated")
    except Exception as e:
        logger.error(f"Admin seed failed: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
